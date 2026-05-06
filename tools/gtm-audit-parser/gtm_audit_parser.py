"""
GTM Audit Parser — Brandalize Hub
Genera automaticamente il Piano di Misurazione da un container GTM esportato.

Uso:
    python gtm_audit_parser.py <percorso_container.json> [nome_cliente]
"""

import json
import sys
import re
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# COSTANTI
# ──────────────────────────────────────────────

GA4_AUTO_EVENTS = {
    "page_view", "session_start", "first_visit", "user_engagement",
    "scroll", "click", "view_search_results", "form_start", "form_submit",
    "video_start", "video_progress", "video_complete", "file_download",
}

TAG_TYPE_LABEL = {
    "gaawe":    "GTM – GA4 Event Tag",
    "googtag":  "GTM – Google Tag (GA4 Config / Ads)",
    "awct":     "GTM – Google Ads Conversion",
    "gclidw":   "GTM – Google Ads Remarketing",
    "img":      "GTM – Pixel Image Tag",
    "html":     "GTM – Custom HTML",
    "cvt_K8GSG":"GTM – Community Template (CMP)",
    "cvt_WV8KV":"GTM – Community Template",
}

TRIGGER_TYPE_LABEL = {
    "PAGEVIEW":      "Visualizzazione pagina",
    "DOM_READY":     "DOM Ready",
    "WINDOW_LOADED": "Window Loaded",
    "CLICK":         "Clic elemento",
    "LINK_CLICK":    "Clic su link",
    "FORM_SUBMIT":   "Submit form",
    "CUSTOM_EVENT":  "Custom Event (dataLayer)",
    "SCROLL_DEPTH":  "Scroll Depth",
    "TIMER":         "Timer",
    "TRIGGER_GROUP": "Gruppo trigger",
    "HISTORY_CHANGE":"Cambio history (SPA)",
    "ELEMENT_VISIBILITY": "Visibilità elemento",
}

# ──────────────────────────────────────────────
# COLORI BRANDALIZE
# ──────────────────────────────────────────────

C_BLACK      = "FF1A1A1A"
C_WHITE      = "FFFFFFFF"
C_HEADER_BG  = "FF1A1A2E"   # blu scuro
C_SUBHEAD_BG = "FF16213E"   # blu medio
C_ACCENT     = "FF0F3460"   # blu accent
C_GA4        = "FFE8F5E9"   # verde pallido
C_META       = "FFFE6F00"   # arancio Meta header
C_META_ROW   = "FFFFF3E0"   # arancio pallido
C_GADS       = "FF4285F4"   # blu Google
C_GADS_ROW   = "FFE8F0FE"   # blu pallido
C_CMP        = "FF6A0DAD"   # viola CMP header
C_CMP_ROW    = "FFF3E5F5"   # viola pallido
C_ALT_ROW    = "FFF8F9FA"   # grigio alternato
C_BORDER     = "FFCCCCCC"
C_OK         = "FF2E7D32"
C_WARN       = "FFF57F17"
C_ERROR      = "FFC62828"

THIN_BORDER = Border(
    left=Side(style="thin", color=C_BORDER),
    right=Side(style="thin", color=C_BORDER),
    top=Side(style="thin", color=C_BORDER),
    bottom=Side(style="thin", color=C_BORDER),
)

# ──────────────────────────────────────────────
# HELPERS STILE
# ──────────────────────────────────────────────

def header_style(cell, bg=C_HEADER_BG, fg=C_WHITE, bold=True, size=10):
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER

def body_style(cell, bg=C_WHITE, wrap=True):
    cell.font = Font(name="Arial", size=9, color=C_BLACK)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
    cell.border = THIN_BORDER

def status_badge(cell, value):
    """Colora la cella in base al valore."""
    body_style(cell)
    v = str(value).lower()
    if v in ("tracciato", "✅ tracciato", "ok"):
        cell.font = Font(name="Arial", size=9, bold=True, color=C_OK)
    elif "attenzione" in v or "warn" in v:
        cell.font = Font(name="Arial", size=9, bold=True, color=C_WARN)
    else:
        cell.font = Font(name="Arial", size=9, color="FF666666")

def write_row(ws, row_num, values, bg=C_WHITE):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        body_style(c, bg=bg)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze_header(ws, row=2):
    ws.freeze_panes = ws.cell(row=row + 1, column=1)

# ──────────────────────────────────────────────
# PARSING GTM
# ──────────────────────────────────────────────

def load_gtm(path):
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    cv = data.get("containerVersion", data)
    return cv

def build_trigger_map(cv):
    triggers = cv.get("trigger", [])
    return {t["triggerId"]: t for t in triggers}

def extract_url_conditions(trigger):
    """Estrae condizioni URL/pagina da un trigger."""
    pages = []
    for key in ("filter", "customEventFilter", "autoEventFilter"):
        for f in trigger.get(key, []):
            params = f.get("parameter", [])
            arg0 = next((p.get("value","") for p in params if p.get("key")=="arg0"), "")
            arg1 = next((p.get("value","") for p in params if p.get("key")=="arg1"), "")
            op = f.get("type", "")
            if any(x in arg0.lower() for x in ("page url", "page path", "{{page")):
                op_label = {
                    "EQUALS": "=",
                    "CONTAINS": "contiene",
                    "STARTS_WITH": "inizia con",
                    "ENDS_WITH": "finisce con",
                    "MATCHES_CSS": "css",
                    "MATCH_REGEX": "regex",
                }.get(op, op)
                pages.append(f"{op_label} {arg1}")
    return ", ".join(pages) if pages else "Tutte le pagine"

def get_trigger_info(trigger_ids, trigger_map):
    """Restituisce (nomi trigger, tipo trigger, pagine)."""
    names, types, pages = [], [], []
    for tid in trigger_ids:
        t = trigger_map.get(tid)
        if not t:
            continue
        names.append(t.get("name", tid))
        ttype = t.get("type", "")
        types.append(TRIGGER_TYPE_LABEL.get(ttype, ttype))
        pages.append(extract_url_conditions(t))
    return (
        " | ".join(names) if names else "–",
        " | ".join(set(types)) if types else "–",
        " | ".join(dict.fromkeys(pages)) if pages else "–",
    )

def extract_ga4_params(tag_params):
    """Estrae parametri evento GA4."""
    dims = []
    for p in tag_params:
        if p.get("key") == "eventSettingsTable":
            for item in p.get("list", []):
                mp = item.get("map", [])
                pname = next((x.get("value","") for x in mp if x.get("key")=="parameter"), "")
                pval  = next((x.get("value","") for x in mp if x.get("key")=="parameterValue"), "")
                if pname:
                    dims.append(f"{pname}: {pval}" if pval else pname)
    return "\n".join(dims) if dims else "–"

def extract_meta_event_from_html(html):
    """Estrae nome evento fbq dalla stringa HTML."""
    patterns = [
        r"fbq\(['\"]track['\"],\s*['\"]([^'\"]+)['\"]",
        r"fbq\(['\"]trackCustom['\"],\s*['\"]([^'\"]+)['\"]",
        r"fbq\(['\"]trackSingle['\"],\s*['\"][^'\"]+['\"],\s*['\"]([^'\"]+)['\"]",
    ]
    for pat in patterns:
        m = re.search(pat, html)
        if m:
            return m.group(1)
    return "–"

def extract_meta_params_from_html(html):
    """Estrae parametri passati all'evento fbq."""
    m = re.search(r"fbq\([^)]*,\s*\{([^}]+)\}", html)
    if m:
        return m.group(1).strip().replace("\n"," ")
    return "–"

def is_custom_event(tag_type, event_name):
    if tag_type == "gaawe":
        return "Automatico" if event_name in GA4_AUTO_EVENTS else "Personalizzato"
    return "–"

# ──────────────────────────────────────────────
# SEZIONI AUDIT
# ──────────────────────────────────────────────

def parse_ga4(cv, trigger_map):
    """Restituisce lista di dict con dati GA4."""
    rows = []
    tags = cv.get("tag", [])
    
    # GA4 Config tag
    for tag in tags:
        if tag.get("type") == "googtag":
            name = tag.get("name","")
            params = tag.get("parameter", [])
            tag_id = next((p.get("value","") for p in params if p.get("key")=="tagId"), "–")
            firing = tag.get("firingTriggerId", [])
            trig_name, trig_type, pages = get_trigger_info(firing, trigger_map)
            rows.append({
                "nome_evento": "page_view (config)",
                "tag_name": name,
                "trigger_nome": trig_name,
                "trigger_tipo": trig_type,
                "come_tracciato": "GTM – Google Tag (GA4 Config)",
                "pagina": pages,
                "parametri": f"Measurement ID: {tag_id}",
                "auto_custom": "Automatico",
                "note": "",
            })
    
    # GA4 Event tags
    for tag in tags:
        if tag.get("type") != "gaawe":
            continue
        name = tag.get("name", "")
        params = tag.get("parameter", [])
        event_name = next((p.get("value","") for p in params if p.get("key")=="eventName"), "–")
        firing = tag.get("firingTriggerId", [])
        trig_name, trig_type, pages = get_trigger_info(firing, trigger_map)
        dims = extract_ga4_params(params)
        rows.append({
            "nome_evento": event_name,
            "tag_name": name,
            "trigger_nome": trig_name,
            "trigger_tipo": trig_type,
            "come_tracciato": "GTM – GA4 Event Tag",
            "pagina": pages,
            "parametri": dims,
            "auto_custom": is_custom_event("gaawe", event_name),
            "note": "",
        })
    return rows

def parse_meta(cv, trigger_map):
    rows = []
    tags = cv.get("tag", [])
    for tag in tags:
        if tag.get("type") != "html":
            continue
        name = tag.get("name", "")
        params = tag.get("parameter", [])
        html_val = next((p.get("value","") for p in params if p.get("key")=="html"), "")
        
        # Solo tag con fbq
        if "fbq" not in html_val:
            continue
        
        event_name = extract_meta_event_from_html(html_val)
        meta_params = extract_meta_params_from_html(html_val)
        
        # Pixel ID
        pixel_id_m = re.search(r"PIXEL_ID\s*=\s*['\"]([^'\"]+)['\"]", html_val)
        if not pixel_id_m:
            pixel_id_m = re.search(r"fbq\('init',\s*'([^']+)'", html_val)
        pixel_id = pixel_id_m.group(1) if pixel_id_m else "–"
        
        # EventID (deduplication)
        has_eventid = "eventID" in html_val or "event_id" in html_val.lower()
        
        firing = tag.get("firingTriggerId", [])
        trig_name, trig_type, pages = get_trigger_info(firing, trigger_map)
        
        rows.append({
            "nome_evento": event_name,
            "tag_name": name,
            "trigger_nome": trig_name,
            "trigger_tipo": trig_type,
            "come_tracciato": "GTM – Custom HTML",
            "pagina": pages,
            "parametri": meta_params,
            "pixel_id": pixel_id,
            "dedup_eventid": "✅ Sì" if has_eventid else "⚠️ No",
            "note": "",
        })
    return rows

def parse_gads(cv, trigger_map):
    rows = []
    tags = cv.get("tag", [])
    for tag in tags:
        if tag.get("type") not in ("awct", "gclidw"):
            continue
        name = tag.get("name", "")
        params = tag.get("parameter", [])
        conv_id    = next((p.get("value","") for p in params if p.get("key")=="conversionId"), "–")
        conv_label = next((p.get("value","") for p in params if p.get("key")=="conversionLabel"), "–")
        value_raw  = next((p.get("value","") for p in params if p.get("key")=="value"), "–")
        currency   = next((p.get("value","") for p in params if p.get("key")=="currencyCode"), "–")
        
        firing = tag.get("firingTriggerId", [])
        trig_name, trig_type, pages = get_trigger_info(firing, trigger_map)
        
        rows.append({
            "nome_conversione": name,
            "conv_id": conv_id,
            "etichetta": conv_label,
            "trigger_nome": trig_name,
            "trigger_tipo": trig_type,
            "come_tracciato": TAG_TYPE_LABEL.get(tag.get("type",""), "GTM"),
            "pagina": pages,
            "valore": value_raw,
            "valuta": currency,
            "note": "",
        })
    return rows

# ──────────────────────────────────────────────
# FOGLI EXCEL
# ──────────────────────────────────────────────

def add_title_row(ws, title, col_span, bg=C_HEADER_BG):
    ws.append([title])
    cell = ws.cell(row=ws.max_row, column=1)
    cell.font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=ws.max_row, start_column=1,
                   end_row=ws.max_row, end_column=col_span)
    ws.row_dimensions[ws.max_row].height = 28

def add_section_header(ws, cols, bg=C_SUBHEAD_BG):
    ws.append(cols)
    r = ws.max_row
    ws.row_dimensions[r].height = 22
    for col in range(1, len(cols)+1):
        header_style(ws.cell(row=r, column=col), bg=bg)

def build_overview(ws, meta, ga4_rows, meta_rows, gads_rows):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 50

    add_title_row(ws, f"📊 AUDIT TRACCIAMENTO — {meta['cliente']}", 3)
    ws.append([])

    info = [
        ("Container GTM", meta["container_id"], ""),
        ("Data audit", meta["data"], ""),
        ("Tag totali", str(meta["tags_tot"]), ""),
        ("Trigger totali", str(meta["triggers_tot"]), ""),
        ("Variabili totali", str(meta["vars_tot"]), ""),
    ]
    add_section_header(ws, ["Informazioni Container", "Valore", "Note"], bg=C_ACCENT)
    for row in info:
        ws.append(list(row))
        r = ws.max_row
        ws.row_dimensions[r].height = 18
        for col in range(1, 4):
            bg = C_ALT_ROW if r % 2 == 0 else C_WHITE
            body_style(ws.cell(row=r, column=col), bg=bg)

    ws.append([])
    summary = [
        ("GA4 – Event tag GTM", len([x for x in ga4_rows if "Config" not in x["come_tracciato"]]), ""),
        ("GA4 – Config tag", len([x for x in ga4_rows if "Config" in x["come_tracciato"]]), ""),
        ("Meta Pixel – eventi tracciati", len(meta_rows), ""),
        ("Meta – eventi con EventID (dedup)", len([x for x in meta_rows if "Sì" in x.get("dedup_eventid","")]), ""),
        ("Google Ads – conversion tag", len(gads_rows), ""),
    ]
    add_section_header(ws, ["Riepilogo Tracciamento", "Conteggio", "Note"], bg=C_ACCENT)
    for row in summary:
        ws.append(list(row))
        r = ws.max_row
        ws.row_dimensions[r].height = 18
        for col in range(1, 4):
            bg = C_ALT_ROW if r % 2 == 0 else C_WHITE
            body_style(ws.cell(row=r, column=col), bg=bg)


def build_ga4_sheet(ws, rows):
    ws.sheet_view.showGridLines = False
    COLS = [
        "Nome Evento", "Tag GTM", "Come viene triggerato",
        "Tipo Trigger", "Come viene tracciato",
        "Pagina / URL", "Parametri / Dimensioni",
        "Auto / Personalizzato", "Note"
    ]
    WIDTHS = [28, 35, 35, 24, 28, 38, 40, 20, 35]

    add_title_row(ws, "GA4 — Mappatura eventi tracciati", len(COLS), bg=C_HEADER_BG)
    add_section_header(ws, COLS, bg="FF2E7D32")
    set_col_widths(ws, WIDTHS)

    for i, row in enumerate(rows):
        bg = C_GA4 if i % 2 == 0 else C_WHITE
        values = [
            row["nome_evento"],
            row["tag_name"],
            row["trigger_nome"],
            row["trigger_tipo"],
            row["come_tracciato"],
            row["pagina"],
            row["parametri"],
            row["auto_custom"],
            row["note"],
        ]
        write_row(ws, ws.max_row + 1, values, bg=bg)
        ws.row_dimensions[ws.max_row].height = 40

    freeze_header(ws)


def build_meta_sheet(ws, rows):
    ws.sheet_view.showGridLines = False
    COLS = [
        "Nome Evento Meta", "Tag GTM", "Come viene triggerato",
        "Tipo Trigger", "Come viene tracciato",
        "Pagina / URL", "Parametri passati",
        "Pixel ID", "EventID (dedup)", "Note"
    ]
    WIDTHS = [25, 38, 35, 22, 24, 38, 38, 20, 16, 35]

    add_title_row(ws, "META ADS — Mappatura eventi Pixel", len(COLS), bg=C_META)
    add_section_header(ws, COLS, bg="FFE65100")
    set_col_widths(ws, WIDTHS)

    for i, row in enumerate(rows):
        bg = C_META_ROW if i % 2 == 0 else C_WHITE
        values = [
            row["nome_evento"],
            row["tag_name"],
            row["trigger_nome"],
            row["trigger_tipo"],
            row["come_tracciato"],
            row["pagina"],
            row["parametri"],
            row["pixel_id"],
            row["dedup_eventid"],
            row["note"],
        ]
        write_row(ws, ws.max_row + 1, values, bg=bg)
        ws.row_dimensions[ws.max_row].height = 40

    freeze_header(ws)


def build_gads_sheet(ws, rows):
    ws.sheet_view.showGridLines = False
    COLS = [
        "Nome Conversione", "Conversion ID", "Etichetta",
        "Come viene triggerato", "Tipo Trigger",
        "Come viene tracciato", "Pagina / URL",
        "Valore", "Valuta", "Note"
    ]
    WIDTHS = [38, 22, 28, 35, 22, 24, 38, 22, 10, 35]

    add_title_row(ws, "GOOGLE ADS — Mappatura conversioni", len(COLS), bg=C_GADS)
    add_section_header(ws, COLS, bg="FF1565C0")
    set_col_widths(ws, WIDTHS)

    for i, row in enumerate(rows):
        bg = C_GADS_ROW if i % 2 == 0 else C_WHITE
        values = [
            row["nome_conversione"],
            row["conv_id"],
            row["etichetta"],
            row["trigger_nome"],
            row["trigger_tipo"],
            row["come_tracciato"],
            row["pagina"],
            row["valore"],
            row["valuta"],
            row["note"],
        ]
        write_row(ws, ws.max_row + 1, values, bg=bg)
        ws.row_dimensions[ws.max_row].height = 40

    freeze_header(ws)


def build_cmp_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 50

    add_title_row(ws, "CMP — Configurazione Consent Management Platform", 4, bg=C_CMP)
    ws.append([])

    sections = [
        ("🔍 IDENTIFICAZIONE CMP", [
            ("Provider CMP", "", "", "es. CookieLawInfo, Iubenda, Cookiebot, OneTrust"),
            ("Versione / Piano", "", "", ""),
            ("Lingua banner", "", "", ""),
            ("Dominio dove è attiva", "", "", ""),
        ]),
        ("⚙️ CONFIGURAZIONE CONSENSO", [
            ("Modalità: Opt-in o Opt-out?", "", "", "GDPR richiede opt-in per marketing"),
            ("Granularità categorie cookie", "", "", "es. Necessari / Analytics / Marketing / Preferenze"),
            ("Consent Mode v2 attivo?", "", "", "Obbligatorio per Google Ads / GA4 da marzo 2024"),
            ("Consent Mode: Basic o Advanced?", "", "", "Advanced = dati modellati anche senza consenso"),
            ("Tag GTM consent default impostato?", "", "", "Controlla trigger 'Consent Initialization'"),
        ]),
        ("🔗 INTEGRAZIONE GTM", [
            ("Cookie consent update tracciato su GTM?", "", "", "Trigger su dataLayer 'consent_update'"),
            ("GA4 rispetta il consenso analytics?", "", "", "Verifica tag GA4 con consent check"),
            ("Meta Pixel rispetta il consenso marketing?", "", "", ""),
            ("Google Ads rispetta il consenso marketing?", "", "", ""),
            ("Pixel si attivano senza consenso?", "", "", "⚠️ VIOLAZIONE GDPR se sì"),
        ]),
        ("🍪 COOKIE AUDIT", [
            ("Cookie di terze parti presenti senza consenso?", "", "", "Verifica con DevTools > Application > Cookies"),
            ("Cookie di sessione vs persistenti", "", "", ""),
            ("Durata cookie consenso", "", "", "Raccomandato: max 12 mesi"),
            ("Cookie wall bloccante?", "", "", "Potenzialmente non conforme GDPR"),
        ]),
        ("📋 DOCUMENTI LEGALI", [
            ("Privacy Policy presente e aggiornata?", "", "", ""),
            ("Cookie Policy separata?", "", "", ""),
            ("Link nel banner al documento?", "", "", ""),
            ("Data ultimo aggiornamento documenti", "", "", ""),
        ]),
        ("✅ VALUTAZIONE FINALE", [
            ("Conformità GDPR stimata", "", "", ""),
            ("Problemi critici rilevati", "", "", ""),
            ("Azioni raccomandate", "", "", ""),
            ("Priorità intervento", "", "", "Alta / Media / Bassa"),
        ]),
    ]

    for section_title, items in sections:
        ws.append([])
        ws.append([section_title, "Valore rilevato", "Stato", "Note / Raccomandazione"])
        r = ws.max_row
        ws.row_dimensions[r].height = 20
        for col in range(1, 5):
            header_style(ws.cell(row=r, column=col), bg=C_CMP)
        for i, item in enumerate(items):
            ws.append(list(item))
            r2 = ws.max_row
            ws.row_dimensions[r2].height = 22
            bg = C_CMP_ROW if i % 2 == 0 else C_WHITE
            for col in range(1, 5):
                body_style(ws.cell(row=r2, column=col), bg=bg)

    freeze_header(ws, row=3)


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────


# ──────────────────────────────────────────────
# VARIABILI
# ──────────────────────────────────────────────

VAR_TYPE_LABEL = {
    "c":    "Costante",
    "v":    "Variabile dataLayer",
    "k":    "Cookie first-party",
    "jsm":  "JavaScript personalizzato",
    "smm":  "Lookup Table (tabella valori)",
    "u":    "Parametro URL / Query String",
    "awec": "Google Ads – Enhanced Conversions (dati utente)",
    "gas":  "Impostazioni Google Analytics",
    "j":    "Variabile JavaScript",
    "d":    "Elemento DOM",
    "f":    "HTTP Referrer",
    "r":    "Regex Table",
}

VAR_CATEGORY = {
    "c":    "🔑 ID & Configurazione",
    "v":    "📦 dataLayer",
    "k":    "🍪 Cookie",
    "jsm":  "⚙️ JavaScript Custom",
    "smm":  "🗂️ Lookup / Mapping",
    "u":    "🔗 URL / UTM",
    "awec": "💰 Google Ads Enhanced Conv.",
    "gas":  "📊 GA Settings",
}

C_VAR        = "FF37474F"   # grigio scuro header
C_VAR_ROW    = "FFECEFF1"   # grigio pallido

def extract_var_value(var):
    """Estrae il valore/fonte leggibile dalla variabile."""
    vtype = var.get("type", "")
    params = var.get("parameter", [])
    
    if vtype == "c":
        return next((p.get("value","") for p in params if p.get("key")=="value"), "–")
    
    if vtype == "v":
        dl_name = next((p.get("value","") for p in params if p.get("key")=="name"), "–")
        default = next((p.get("value","") for p in params if p.get("key")=="defaultValue"), None)
        return f"dataLayer.{dl_name}" + (f" (default: {default})" if default else "")
    
    if vtype == "k":
        cookie_name = next((p.get("value","") for p in params if p.get("key")=="name"), "–")
        return f"Cookie: {cookie_name}"
    
    if vtype == "u":
        component = next((p.get("value","") for p in params if p.get("key")=="component"), "")
        query_key = next((p.get("value","") for p in params if p.get("key")=="queryKey"), "")
        if component == "QUERY":
            return f"Query param: ?{query_key}"
        return component
    
    if vtype == "smm":
        input_var = next((p.get("value","") for p in params if p.get("key")=="input"), "–")
        default_v = next((p.get("value","") for p in params if p.get("key")=="defaultValue"), "")
        return f"Input: {input_var}" + (f" → default: {default_v}" if default_v else "")
    
    if vtype == "jsm":
        js = next((p.get("value","") for p in params if p.get("key")=="javascript"), "")
        # Prendi prima riga significativa del codice
        lines = [l.strip() for l in js.splitlines() if l.strip() and not l.strip().startswith("//") and l.strip() != "function() {" and l.strip() != "}"]
        return lines[0][:120] if lines else "JS custom"
    
    if vtype == "awec":
        mode = next((p.get("value","") for p in params if p.get("key")=="mode"), "")
        email = next((p.get("value","") for p in params if p.get("key")=="email"), "")
        return f"Modalità: {mode}" + (f" | Email: {email}" if email else "")
    
    return "–"

def describe_variable(var):
    """Genera una descrizione leggibile di cosa fa la variabile."""
    vtype = var.get("type", "")
    name = var.get("name", "")
    
    descriptions = {
        # IDs
        "GA4 ID": "Measurement ID di GA4, usato nel tag di configurazione Google Tag",
        "Google Ads ID": "ID account Google Ads, usato come prefisso nelle conversioni",
        "Google Grants ID": "ID account Google Grants (Ads nonprofit), usato nelle conversioni Grants",
        "Event ID": "ID univoco dell'evento generato da community template, usato per deduplicare eventi Meta Pixel lato browser e server (CAPI)",
        # Consent
        "LK - Meta Consent": "Mappa il valore di ad_storage nel formato richiesto da Meta (grant/revoke)",
        "ad_storage": "Stato del consenso per cookie pubblicitari (Consent Mode v2)",
        "analytics_storage": "Stato del consenso per cookie analytics (Consent Mode v2)",
        # Cookie Meta
        "Cookie - _fbp": "Cookie _fbp di Meta: identificatore browser per attribuzione annunci",
        "Cookie - _fbc": "Cookie _fbc di Meta: contiene fbclid, usato per attribuzione click",
        "fbclid": "Parametro fbclid dall'URL: ID click Meta Ads, usato per costruire _fbc",
        "fbclid present": "Restituisce true/false se fbclid è presente in URL: usato come condizione trigger",
        "fbc - final": "Costruisce il valore fbc definitivo: legge cookie _fbc, altrimenti costruisce da fbclid",
        "FBP - Cookie o Sintetico": "Legge _fbp da cookie o costruisce valore sintetico se assente (utenti senza cookie)",
        # Dati donatore
        "donation_value": "Valore della donazione dal dataLayer, passato a Meta/GA4/GAds come parametro value",
        "payment_method": "Metodo di pagamento selezionato dall'utente (carta/PayPal) dal dataLayer",
        "donation_type": "Tipo donazione (singola/regolare) dal dataLayer",
        "donation_id": "ID univoco donazione dal dataLayer, usato per deduplicazione CAPI",
        "currency": "Valuta donazione (EUR) dal dataLayer",
        "donation_frequency": "Frequenza donazione dal dataLayer",
        "donor_data.email": "Email del donatore dal dataLayer – PII, usata per Enhanced Conversions",
        "donor_data.first_name": "Nome del donatore dal dataLayer – PII",
        "donor_data.last_name": "Cognome del donatore dal dataLayer – PII",
        "am_params": "Oggetto parametri Advanced Matching per Meta Pixel, costruito prima dell'init",
        # Form
        "nf_form_name": "Nome del form Ninja Forms dal dataLayer, passato all'evento form_submit",
        "nome_form": "Nome form mappato: traduce il nome tecnico Ninja Forms in etichetta leggibile",
        "field_name": "Nome del campo form compilato (usato per tracciare email durante donazione)",
        # UTM
        "utm_campaign": "Parametro UTM campaign dall'URL della sessione corrente",
        "utm_content": "Parametro UTM content dall'URL della sessione corrente",
        "LS - utm_campaign": "UTM campaign salvato in localStorage (persiste tra pagine senza UTM)",
        "LS - utm_content": "UTM content salvato in localStorage",
        # Altri
        "Timestamp": "Unix timestamp in millisecondi, usato per costruire _fbc e EventID",
        "userAgent": "User Agent del browser, inviato a Meta CAPI per corrispondenza server-client",
        "Dati forniti dagli utenti": "Oggetto Enhanced Conversions Google Ads: contiene email hashata del donatore",
        "eventModel.file_name": "Nome file scaricato, dal modello evento GA4 enhanced measurement",
        "JS - recurring_donation": "Legge il campo recurring_donation dal DOM e restituisce 'Una volta' o 'Regolarmente'",
        "Charitable - Tipo Donazione Raw": "Legge il valore raw del campo recurring_donation dal DOM (once/recurring)",
    }
    
    if name in descriptions:
        return descriptions[name]
    
    # Fallback per tipo
    if vtype == "v":
        params = var.get("parameter", [])
        dl_name = next((p.get("value","") for p in params if p.get("key")=="name"), "")
        return f"Legge il valore '{dl_name}' dal dataLayer"
    if vtype == "k":
        params = var.get("parameter", [])
        cn = next((p.get("value","") for p in params if p.get("key")=="name"), "")
        return f"Legge il cookie '{cn}' dal browser"
    if vtype == "u":
        params = var.get("parameter", [])
        qk = next((p.get("value","") for p in params if p.get("key")=="queryKey"), "")
        return f"Legge il parametro URL '{qk}'"
    if vtype == "c":
        return "Valore costante usato come riferimento nei tag"
    if vtype == "smm":
        return "Mappa un valore in input a un valore di output tramite tabella"
    if vtype == "jsm":
        return "Script JavaScript personalizzato che calcola/restituisce un valore a runtime"
    
    return "–"

def find_var_usages(var_name, cv):
    """Trova i tag che usano questa variabile (cerca {{var_name}} nel JSON serializzato dei tag)."""
    search_str = f"{{{{{var_name}}}}}"
    used_in = []
    for tag in cv.get("tag", []):
        tag_str = json.dumps(tag)
        if search_str in tag_str:
            used_in.append(tag.get("name", "?"))
    return ", ".join(used_in) if used_in else "–"

def parse_variables(cv):
    rows = []
    for var in cv.get("variable", []):
        vtype = var.get("type", "")
        name = var.get("name", "")
        rows.append({
            "nome": name,
            "tipo": VAR_TYPE_LABEL.get(vtype, vtype),
            "categoria": VAR_CATEGORY.get(vtype, "🔧 Altro"),
            "descrizione": describe_variable(var),
            "valore_fonte": extract_var_value(var),
            "usata_in": find_var_usages(name, cv),
        })
    # Ordina per categoria poi nome
    rows.sort(key=lambda x: (x["categoria"], x["nome"]))
    return rows

def build_variables_sheet(ws, rows):
    ws.sheet_view.showGridLines = False
    COLS = [
        "Nome Variabile", "Categoria", "Tipo GTM",
        "Descrizione – Cosa fa", "Valore / Fonte",
        "Usata nei tag"
    ]
    WIDTHS = [32, 26, 30, 55, 45, 55]

    add_title_row(ws, "VARIABILI GTM — Dettaglio completo", len(COLS), bg=C_VAR)
    add_section_header(ws, COLS, bg="FF263238")
    set_col_widths(ws, WIDTHS)

    current_cat = None
    for i, row in enumerate(rows):
        # Riga separatore per categoria
        if row["categoria"] != current_cat:
            current_cat = row["categoria"]
            ws.append([current_cat, "", "", "", "", ""])
            r = ws.max_row
            ws.row_dimensions[r].height = 18
            for col in range(1, 7):
                c = ws.cell(row=r, column=col)
                c.font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
                c.fill = PatternFill("solid", fgColor="FF455A64")
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = THIN_BORDER

        bg = C_VAR_ROW if i % 2 == 0 else C_WHITE
        values = [
            row["nome"],
            row["categoria"],
            row["tipo"],
            row["descrizione"],
            row["valore_fonte"],
            row["usata_in"],
        ]
        write_row(ws, ws.max_row + 1, values, bg=bg)
        ws.row_dimensions[ws.max_row].height = 42

    freeze_header(ws)


def run(json_path, cliente="Cliente"):
    cv = load_gtm(json_path)
    trigger_map = build_trigger_map(cv)

    container_info = cv.get("container", {})
    container_id = container_info.get("publicId", "–")
    tags = cv.get("tag", [])
    triggers = cv.get("trigger", [])
    variables = cv.get("variable", [])

    ga4_rows  = parse_ga4(cv, trigger_map)
    meta_rows = parse_meta(cv, trigger_map)
    gads_rows = parse_gads(cv, trigger_map)
    var_rows  = parse_variables(cv)

    meta = {
        "cliente": cliente,
        "container_id": container_id,
        "data": datetime.now().strftime("%d/%m/%Y"),
        "tags_tot": len(tags),
        "triggers_tot": len(triggers),
        "vars_tot": len(variables),
    }

    wb = Workbook()

    # Rinomina foglio default
    wb.active.title = "Overview"
    build_overview(wb.active, meta, ga4_rows, meta_rows, gads_rows)

    ws_ga4 = wb.create_sheet("GA4")
    build_ga4_sheet(ws_ga4, ga4_rows)

    ws_meta = wb.create_sheet("Meta Ads")
    build_meta_sheet(ws_meta, meta_rows)

    ws_gads = wb.create_sheet("Google Ads")
    build_gads_sheet(ws_gads, gads_rows)

    ws_cmp = wb.create_sheet("CMP")
    build_cmp_sheet(ws_cmp)

    ws_vars = wb.create_sheet("Variabili GTM")
    build_variables_sheet(ws_vars, var_rows)

    # Output — salva nella stessa cartella del file JSON
    out_name = f"Audit_Tracciamento_{cliente.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    out_path = os.path.join(os.path.dirname(os.path.abspath(json_path)), out_name)
    wb.save(out_path)
    print(f"✅ File salvato: {out_path}")
    print(f"   GA4 eventi:       {len(ga4_rows)}")
    print(f"   Meta eventi:      {len(meta_rows)}")
    print(f"   Google Ads conv.: {len(gads_rows)}")
    print(f"   Variabili:        {len(var_rows)}")
    return out_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python gtm_audit_parser.py <container.json> [nome_cliente]")
        sys.exit(1)
    json_file = sys.argv[1]
    client_name = sys.argv[2] if len(sys.argv) > 2 else "Cliente"
    run(json_file, client_name)
